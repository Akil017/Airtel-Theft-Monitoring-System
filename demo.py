"""
Airtel BTS Theft Monitoring System — IRON MAN HUD Edition
==========================================================
Restricted Area Intrusion Detection using YOLOv8

  Severity:  1-2 persons → HIGH  |  3+ persons → CRITICAL
  Model:     YOLOv8s (higher accuracy than nano)
  HUD:       Tactical overlay
"""

import cv2
import math
import time
import os
import threading
import sys
import numpy as np
from datetime import datetime
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
#  SOUND
# ═══════════════════════════════════════════════════════════════════════════════
if sys.platform == "win32":
    import winsound
    def play_alarm(severity="HIGH"):
        patterns = {
            "HIGH":     [(1100, 200), (0, 60), (1300, 200), (0, 60), (1100, 300)],
            "CRITICAL": [(1600, 120), (0, 40)] * 8 + [(2000, 500)],
        }
        def _beep():
            for freq, dur in patterns.get(severity, patterns["HIGH"]):
                if freq > 0: winsound.Beep(freq, dur)
                else:        time.sleep(dur / 1000)
        threading.Thread(target=_beep, daemon=True).start()
else:
    def play_alarm(severity="HIGH"):
        reps = 8 if severity == "CRITICAL" else 4
        for _ in range(reps):
            print("\a", end="", flush=True)
            time.sleep(0.08)

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
CAMERA_SOURCE        = 0
CAMERA_ID            = "CAM-BTS-01"
SITE_ID              = "AIRTEL-ASM-BTS-001"
SITE_NAME            = "Guwahati North BTS"
SITE_LOCATION        = "Guwahati, Assam"
ZONE                 = "RESTRICTED ZONE — AUTHORISED PERSONNEL ONLY"
OPERATOR             = "NOC-OPS-ASM-01"
MODEL_PATH           = "yolov8s.pt"          # 's' model = better accuracy than nano
CONFIDENCE_THRESHOLD = 0.45
NMS_IOU_THRESHOLD    = 0.40                  # tighter NMS = fewer duplicate boxes
CONSECUTIVE_REQUIRED = 3
ALARM_COOLDOWN_SEC   = 10
LOG_FILE             = "demo_alarms.xlsx"
SNAPSHOT_DIR         = "snapshots"

ANIMAL_CLASSES = {
    15:"Cat", 16:"Dog", 17:"Horse", 18:"Sheep",
    19:"Cow", 20:"Elephant", 21:"Bear", 22:"Zebra", 23:"Giraffe",
}

# ═══════════════════════════════════════════════════════════════════════════════
#  SEVERITY
# ═══════════════════════════════════════════════════════════════════════════════
def get_severity(n):
    if n == 0:
        return "CLEAR"
    elif n == 1:
        return "MID-HIGH"
    elif n == 2:
        return "HIGH"
    else:
        return "CRITICAL"
def get_threat(n):      return "MASS INTRUSION" if n >= 3 else "COORDINATED INTRUSION" if n == 2 else "SINGLE INTRUDER"
def get_response(n):
    if n >= 3: return "DISPATCH MULTIPLE UNITS — MASS INTRUSION"
    if n == 2: return "DISPATCH SECURITY — COORDINATED ENTRY"
    return "DISPATCH SECURITY — SINGLE INTRUDER"

# ═══════════════════════════════════════════════════════════════════════════════
#  COLOUR PALETTE  (BGR)
# ═══════════════════════════════════════════════════════════════════════════════
C = {
    "bg":        (8,   8,   8  ),
    "panel":     (14,  14,  14 ),
    "panel2":    (20,  20,  20 ),
    "border":    (40,  40,  40 ),
    "border2":   (60,  60,  60 ),
    "cyan":      (220, 200, 0  ),   # HUD accent — cyan in BGR
    "cyan_dim":  (120, 110, 0  ),
    "cyan_glow": (255, 240, 60 ),
    "white":     (255, 255, 255),
    "gray":      (120, 120, 120),
    "gray_dim":  (60,  60,  60 ),
    "green":     (60,  210, 60 ),
    "green_dim": (30,  100, 30 ),
    "orange":    (0,   160, 255),
    "orange_dim":(0,   80,  140),
    "red":       (30,  30,  220),
    "red_dim":   (20,  20,  100),
    "critical":  (10,  10,  200),
    "airtel":    (30,  30,  220),
}

def sev_col(sev):
    return C["critical"] if sev == "CRITICAL" else C["red"]

# ═══════════════════════════════════════════════════════════════════════════════
#  DRAWING PRIMITIVES
# ═══════════════════════════════════════════════════════════════════════════════
def put(frame, text, x, y, scale=0.42, color=None, thick=1, font=cv2.FONT_HERSHEY_SIMPLEX):
    cv2.putText(frame, text, (x, y), font, scale, color or C["white"], thick, cv2.LINE_AA)

def put_c(frame, text, cx, y, scale=0.42, color=None, thick=1):
    (tw, _), _ = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, scale, thick)
    put(frame, text, cx - tw // 2, y, scale, color, thick)

def line(frame, p1, p2, col, thick=1):
    cv2.line(frame, p1, p2, col, thick, cv2.LINE_AA)

def rect(frame, p1, p2, col, thick=-1):
    cv2.rectangle(frame, p1, p2, col, thick)

def circle(frame, c, r, col, thick=-1):
    cv2.circle(frame, c, r, col, thick, cv2.LINE_AA)

def alpha_rect(frame, x1, y1, x2, y2, color, alpha=0.55):
    x1,y1,x2,y2 = int(x1),int(y1),int(x2),int(y2)
    """Semi-transparent filled rectangle."""
    overlay = frame.copy()
    cv2.rectangle(overlay, (x1, y1), (x2, y2), color, -1)
    cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)

def corner_bracket(frame, x1, y1, x2, y2, col, L=18, thick=2):
    x1,y1,x2,y2 = int(x1),int(y1),int(x2),int(y2)
    """Iron Man style corner brackets around a bounding box."""
    pts = [(x1,y1,1,1),(x2,y1,-1,1),(x1,y2,1,-1),(x2,y2,-1,-1)]
    for px, py, dx, dy in pts:
        line(frame, (px, py), (px + dx*L, py), col, thick)
        line(frame, (px, py), (px, py + dy*L), col, thick)

def draw_bar(frame, x, y, w, h, val, max_val, fg, bg=(30,30,30)):
    """Horizontal progress bar."""
    rect(frame, (x, y), (x+w, y+h), bg)
    fill = int(w * min(val, max_val) / max_val)
    if fill > 0:
        rect(frame, (x, y), (x+fill, y+h), fg)
    rect(frame, (x, y), (x+w, y+h), C["border"], 1)

def draw_arc_ring(frame, cx, cy, r, start_deg, end_deg, col, thick=2):
    """Partial circle arc — used for radar rings."""
    cv2.ellipse(frame, (cx, cy), (r, r), 0, start_deg, end_deg, col, thick, cv2.LINE_AA)

def scanline_overlay(frame, alpha=0.06):
    """Subtle horizontal scan-line CRT effect."""
    h, w = frame.shape[:2]
    mask = np.zeros((h, w, 3), dtype=np.uint8)
    mask[::3] = 18
    cv2.addWeighted(frame, 1.0, mask, alpha, 0, frame)

def hex_pattern(frame, x1, y1, x2, y2, col=(30,30,30), step=28):
    """Faint hexagonal grid overlay on a region."""
    for row in range(y1, y2, step):
        offset = (step // 2) if ((row - y1) // step) % 2 else 0
        for col_x in range(x1 + offset, x2, step):
            cv2.drawMarker(frame, (col_x, row), col,
                           cv2.MARKER_DIAMOND, 4, 1, cv2.LINE_AA)

# ═══════════════════════════════════════════════════════════════════════════════
#  DETECTION BOX  — Iron Man target lock
# ═══════════════════════════════════════════════════════════════════════════════
def draw_target_box(frame, x1, y1, x2, y2, conf, idx, severity, det_label, t):
    is_person = det_label == "Person"
    base_col  = sev_col(severity) if is_person else C["orange"]
    dim_col   = C["red_dim"]      if is_person else C["orange_dim"]

    bw, bh = x2 - x1, y2 - y1

    # Semi-transparent fill
    alpha_rect(frame, int(x1), int(y1), int(x2), int(y2), dim_col, 0.18)

    # Main bounding rect (thin)
    rect(frame, (int(x1), int(y1)), (int(x2), int(y2)), base_col, 1)

    # Corner brackets (thick)
    corner_bracket(frame, int(x1), int(y1), int(x2), int(y2), base_col, L=20, thick=2)

    # Animated lock-on pulse ring on center
    cx, cy = int((x1+x2)//2), int((y1+y2)//2)
    pulse_r = int(18 + 4 * math.sin(t * 3.0 + idx))
    cv2.ellipse(frame, (cx, cy), (pulse_r, pulse_r), 0, 0, 360,
                base_col, 1, cv2.LINE_AA)
    circle(frame, (cx, cy), 3, base_col)

    # Crosshair on center
    ch = 10
    line(frame, (cx - ch, cy), (cx - 4, cy), base_col)
    line(frame, (cx + 4,  cy), (cx + ch, cy), base_col)
    line(frame, (cx, cy - ch), (cx, cy - 4), base_col)
    line(frame, (cx, cy + 4 ), (cx, cy + ch), base_col)

    # Target ID top-left
    tag = f"TGT-{idx:02d}  {det_label.upper()}"
    (tw, th), _ = cv2.getTextSize(tag, cv2.FONT_HERSHEY_SIMPLEX, 0.40, 1)
    alpha_rect(frame, x1, y1 - th - 14, x1 + tw + 12, y1, C["bg"], 0.80)
    put(frame, tag, int(x1 + 4), int(y1 - 4), 0.40, base_col)

    # Confidence bar (bottom of box)
    bar_y = y2 - 12
    if bar_y > y1 + 10:
        draw_bar(frame, int(x1 + 4), int(bar_y), int(bw - 8), 6, conf, 1.0, base_col, (20,20,20))

    # Conf % bottom-right
    conf_text = f"{conf:.0%}"
    (ct, _), _ = cv2.getTextSize(conf_text, cv2.FONT_HERSHEY_SIMPLEX, 0.38, 1)
    put(frame, conf_text, int(x2 - ct - 4), int(y2 + 14), 0.38, base_col)

    # Distance estimation (fake — based on box height as proxy)
    est_dist = max(1.0, round(180 / max(bh, 1), 1))
    put(frame, f"~{est_dist}m", int(x1 + 4), int(y2 + 14), 0.36, C["gray"])

    # Corner index badge
    circle(frame, (x1 + 12, y1 + 12), 10, base_col)
    badge_x = int(x1 + 8) if idx < 10 else int(x1 + 5)
    put(frame, str(idx), badge_x, int(y1 + 17),
        0.42, C["bg"], 1, cv2.FONT_HERSHEY_DUPLEX)

# ═══════════════════════════════════════════════════════════════════════════════
#  HUD
# ═══════════════════════════════════════════════════════════════════════════════
def draw_hud(frame, fps, consecutive, alarming, alarm_count,
             person_count, best_conf, severity, alarm_start,
             last_cleared_ts, detections, session_start, t):

    h, w = frame.shape[:2]
    PANEL_W = 260
    px = w - PANEL_W

    # ── Scan-line effect ──────────────────────────────────────────────────────
    scanline_overlay(frame, 0.08)

    # ── Hex pattern on panel background ──────────────────────────────────────
    hex_pattern(frame, px, 84, w, h - 58, (22, 22, 22), 28)

    # ══════════════════════════════════════════════════════════════════════════
    #  TOP BAR
    # ══════════════════════════════════════════════════════════════════════════
    rect(frame, (0, 0), (w, 64), C["bg"])
    # Gradient-like layers
    rect(frame, (0, 0), (w, 2), C["cyan"])
    line(frame, (0, 64), (w, 64), C["border2"])

    # Airtel logo with glow effect
    cv2.putText(frame, "AIRTEL", (14, 26),
                cv2.FONT_HERSHEY_DUPLEX, 0.70, C["airtel"], 2, cv2.LINE_AA)
    cv2.putText(frame, "AIRTEL", (14, 26),
                cv2.FONT_HERSHEY_DUPLEX, 0.70, (80, 80, 255), 1, cv2.LINE_AA)

    # Separator dot
    circle(frame, (95, 20), 3, C["border2"])

    # System title
    put(frame, "BTS TACTICAL SURVEILLANCE SYSTEM", 104, 26,
        0.52, C["white"], 1, cv2.FONT_HERSHEY_DUPLEX)

    # Sub info
    put(frame, f"{SITE_NAME}  ·  {SITE_ID}  ·  {SITE_LOCATION}", 14, 52,
        0.36, C["gray"])

    # Uptime
    uptime = int(time.time() - session_start)
    h_up, m_up, s_up = uptime // 3600, (uptime % 3600) // 60, uptime % 60
    put(frame, f"UPTIME  {h_up:02d}:{m_up:02d}:{s_up:02d}", w - 200, 52,
        0.34, C["cyan_dim"])

    # Live indicator (animated)
    dot_col = C["green"] if not alarming else sev_col(severity)
    if alarming and int(t * 3) % 2 == 0:
        circle(frame, (w - 22, 20), 8, dot_col, -1)
    else:
        circle(frame, (w - 22, 20), 6, dot_col, -1)
        circle(frame, (w - 22, 20), 9, dot_col, 1)
    put(frame, "LIVE" if not alarming else "ALARM", w - 80, 26,
        0.40, dot_col, 1)
    put(frame, datetime.now().strftime("%d %b %Y  %H:%M:%S"),
        w - 210, 52, 0.34, C["gray"])

    # ══════════════════════════════════════════════════════════════════════════
    #  ZONE RIBBON
    # ══════════════════════════════════════════════════════════════════════════
    ribbon_col  = (0, 0, 100) if alarming else (10, 40, 10)
    ribbon_text = (70, 70, 255) if alarming else (60, 180, 60)
    flash = alarming and int(t * 4) % 2 == 0
    if flash:
        ribbon_col = (20, 20, 160)
    rect(frame, (0, 65), (w, 84), ribbon_col)
    zone_text = f"  ⚠  {ZONE}  ⚠" if alarming else f"  ▶  {ZONE}"
    put(frame, zone_text, 10, 79, 0.38, ribbon_text)

    # ══════════════════════════════════════════════════════════════════════════
    #  RIGHT PANEL  (JARVIS stats)
    # ══════════════════════════════════════════════════════════════════════════
    rect(frame, (px - 12, 84), (w, h - 58), C["panel"])
    line(frame, (px - 12, 84), (px - 12, h - 58), C["cyan_dim"])
    # Panel top accent line
    line(frame, (px - 12, 84), (w, 84), C["cyan_dim"])

    # Panel title
    rect(frame, (px - 10, 84), (w, 106), C["panel2"])
    put_c(frame, "◈  TACTICAL STATUS  ◈", px + PANEL_W // 2 - 12, 100,
          0.38, C["cyan"])

    # ── Stat blocks ──────────────────────────────────────────────────────────
    def stat_block(label, val, y, vcol, bar_val=None, bar_max=1.0):
        rect(frame, (px - 8, y), (w - 2, y + 38), C["panel2"])
        line(frame, (px - 8, y + 38), (w - 2, y + 38), C["border"])
        # Label
        put(frame, label, px, y + 12, 0.30, C["gray"])
        # Value
        put(frame, str(val), px, y + 30, 0.48, vcol, 1)
        # Optional bar
        if bar_val is not None:
            draw_bar(frame, w - 70, y + 14, 60, 6,
                     bar_val, bar_max, vcol, C["border"])

    sy = 108
    stat_block("PERSONS IN FRAME",
               f"{person_count}  DETECTED" if person_count else "0  CLEAR",
               sy, C["red"] if person_count else C["green"],
               person_count, 5)

    stat_block("THREAT LEVEL",
               get_threat(person_count) if person_count else "NONE",
               sy + 40, sev_col(severity) if person_count else C["green"])

    stat_block("BEST CONFIDENCE",
               f"{best_conf*100:.1f}%" if best_conf else "—",
               sy + 80, C["orange"] if best_conf else C["gray"],
               best_conf, 1.0)

    stat_block("SEVERITY",
               severity if alarming else "CLEAR",
               sy + 120, sev_col(severity) if alarming else C["green"])

    stat_block("ALARM COUNT",
               f"ALM-{alarm_count:04d}",
               sy + 160, C["cyan"])

    # Alarm duration / last clear
    if alarming and alarm_start:
        dur = time.time() - alarm_start
        stat_block("ALARM DURATION",
                   f"{int(dur)}s  ONGOING",
                   sy + 200, sev_col(severity),
                   min(dur, 60), 60)
    elif last_cleared_ts:
        stat_block("LAST CLEARED", last_cleared_ts, sy + 200, C["gray"])
    else:
        stat_block("ALARM DURATION", "—", sy + 200, C["gray"])

    # ── Detection window blocks ───────────────────────────────────────────────
    dw_y = sy + 248
    rect(frame, (px - 8, dw_y), (w - 2, dw_y + 54), C["panel2"])
    put(frame, "DETECTION WINDOW", px, dw_y + 14, 0.30, C["gray"])
    block_w = 34
    spacing = 10
    total_w = CONSECUTIVE_REQUIRED * block_w + (CONSECUTIVE_REQUIRED - 1) * spacing
    bx_start = px + (PANEL_W - total_w) // 2 - 10
    for i in range(CONSECUTIVE_REQUIRED):
        bx = bx_start + i * (block_w + spacing)
        filled = i < consecutive
        if filled and alarming:
            bg = sev_col(severity)
            border_c = sev_col(severity)
        elif filled:
            bg = C["green_dim"]
            border_c = C["green"]
        else:
            bg = C["border"]
            border_c = C["border2"]
        rect(frame, (bx, dw_y + 22), (bx + block_w, dw_y + 46), bg)
        rect(frame, (bx, dw_y + 22), (bx + block_w, dw_y + 46), border_c, 1)
        put_c(frame, str(i + 1), bx + block_w // 2, dw_y + 40, 0.34,
              C["white"] if filled else C["gray_dim"])

    # ── Per-detection mini cards ──────────────────────────────────────────────
    card_y = sy + 310
    put(frame, "TARGET MANIFEST", px, card_y, 0.30, C["gray"])
    card_y += 14
    for di, (conf, x1, y1, x2, y2, lbl) in enumerate(detections[:4]):
        cy_card = card_y + di * 34
        is_p = lbl == "Person"
        tc = sev_col(severity) if is_p else C["orange"]
        rect(frame, (px - 8, cy_card), (w - 2, cy_card + 30), C["panel"])
        rect(frame, (px - 8, cy_card), (px - 4, cy_card + 30), tc, -1)
        put(frame, f"TGT-{di+1:02d}", px + 4, cy_card + 12, 0.32, C["gray"])
        put(frame, lbl.upper(), px + 4, cy_card + 26, 0.38, tc)
        draw_bar(frame, px + 68, cy_card + 10, 80, 6, conf, 1.0, tc)
        put(frame, f"{conf:.0%}", px + 152, cy_card + 26, 0.34, tc)

    # ── Action panel ─────────────────────────────────────────────────────────
    act_y = h - 57
    rect(frame, (px - 8, act_y - 72), (w - 2, act_y), (18, 8, 8))
    line(frame, (px - 8, act_y - 72), (w - 2, act_y - 72), (80, 20, 20))
    put(frame, "REQUIRED ACTION", px, act_y - 58, 0.30, (100, 80, 200))
    if alarming and person_count:
        resp  = get_response(person_count)
        words = resp.split()
        mid   = len(words) // 2
        l1    = " ".join(words[:mid])
        l2    = " ".join(words[mid:])
        put(frame, l1, px, act_y - 40, 0.36, (100, 100, 255))
        put(frame, l2, px, act_y - 22, 0.36, (100, 100, 255))
    else:
        put(frame, "AREA SECURE — MONITORING", px, act_y - 32, 0.36, C["green"])
    put(frame, f"OPR: {OPERATOR}  ·  {CAMERA_ID}", px, act_y - 6,
        0.30, C["gray_dim"])

    # ══════════════════════════════════════════════════════════════════════════
    #  BOTTOM BAR
    # ══════════════════════════════════════════════════════════════════════════
    feed_w = px - 12
    rect(frame, (0, h - 57), (feed_w, h), C["bg"])
    line(frame, (0, h - 57), (feed_w, h - 57), C["border2"])
    # Animated scan line across bottom
    scan_x = int((time.time() % 3) / 3 * feed_w)
    line(frame, (scan_x, h - 57), (min(scan_x + 80, feed_w), h - 57),
         C["cyan_dim"], 2)

    put(frame, f"FPS  {fps:05.1f}", 14, h - 36, 0.36, C["cyan_dim"])
    put(frame, f"MODEL  {MODEL_PATH}", 14, h - 18, 0.34, C["gray_dim"])
    put(frame, f"CONF THRESH  {CONFIDENCE_THRESHOLD:.0%}", 180, h - 36,
        0.34, C["gray_dim"])
    put(frame, f"IOU  {NMS_IOU_THRESHOLD:.0%}", 180, h - 18, 0.34, C["gray_dim"])

    # Mini radar / compass in bottom bar
    radar_cx, radar_cy = feed_w - 70, h - 30
    radar_r = 20
    circle(frame, (radar_cx, radar_cy), radar_r, C["border2"])
    circle(frame, (radar_cx, radar_cy), radar_r // 2, C["border"])
    circle(frame, (radar_cx, radar_cy), 2, C["cyan_dim"])
    line(frame, (radar_cx - radar_r, radar_cy),
         (radar_cx + radar_r, radar_cy), C["border"])
    line(frame, (radar_cx, radar_cy - radar_r),
         (radar_cx, radar_cy + radar_r), C["border"])
    # Rotating sweep line
    angle = (t * 60) % 360
    rx = int(radar_cx + radar_r * math.cos(math.radians(angle)))
    ry = int(radar_cy + radar_r * math.sin(math.radians(angle)))
    line(frame, (radar_cx, radar_cy), (rx, ry), C["cyan"], 1)
    # Blips for each detection
    for di, (conf, x1, y1, x2, y2, _) in enumerate(detections[:4]):
        bx_det = int((x1 + x2) / 2 / frame.shape[1] * radar_r * 2) + radar_cx - radar_r
        by_det = int((y1 + y2) / 2 / frame.shape[0] * radar_r * 2) + radar_cy - radar_r
        bx_det = max(radar_cx - radar_r, min(radar_cx + radar_r, bx_det))
        by_det = max(radar_cy - radar_r, min(radar_cy + radar_r, by_det))
        if int(t * 2) % 2 == 0:
            circle(frame, (bx_det, by_det), 3, C["red"])
    put(frame, "RADAR", radar_cx - 14, radar_cy + radar_r + 12, 0.28, C["gray_dim"])

    # ══════════════════════════════════════════════════════════════════════════
    #  FEED OVERLAY  — corner decorations
    # ══════════════════════════════════════════════════════════════════════════
    # Frame corners
    CORNER_L = 30
    fc = C["cyan_dim"]
    # TL
    line(frame, (0, 85), (CORNER_L, 85), fc)
    line(frame, (0, 85), (0, 85 + CORNER_L), fc)
    # TR (feed area only)
    line(frame, (feed_w, 85), (feed_w - CORNER_L, 85), fc)
    line(frame, (feed_w, 85), (feed_w, 85 + CORNER_L), fc)
    # BL
    line(frame, (0, h - 58), (CORNER_L, h - 58), fc)
    line(frame, (0, h - 58), (0, h - 58 - CORNER_L), fc)
    # BR
    line(frame, (feed_w, h - 58), (feed_w - CORNER_L, h - 58), fc)
    line(frame, (feed_w, h - 58), (feed_w, h - 58 - CORNER_L), fc)

    # Center crosshair on feed
    fh_cx, fh_cy = feed_w // 2, (h - 58 + 85) // 2
    cross_r = 12
    alpha_rect(frame, fh_cx - 1, fh_cy - cross_r,
               fh_cx + 1, fh_cy + cross_r, C["cyan_dim"], 0.5)
    alpha_rect(frame, fh_cx - cross_r, fh_cy - 1,
               fh_cx + cross_r, fh_cy + 1, C["cyan_dim"], 0.5)
    draw_arc_ring(frame, fh_cx, fh_cy, 20, 0, 90, C["cyan_dim"])
    draw_arc_ring(frame, fh_cx, fh_cy, 20, 180, 270, C["cyan_dim"])

    # ══════════════════════════════════════════════════════════════════════════
    #  ALARM OVERLAY  — big flashing banner
    # ══════════════════════════════════════════════════════════════════════════
    if alarming:
        # Pulsing red border on feed area
        if int(t * 3) % 2 == 0:
            rect(frame, (0, 85), (feed_w, h - 58), sev_col(severity), 3)

        # Semi-transparent dark band in middle
        alpha_rect(frame, 0, fh_cy - 52, feed_w, fh_cy + 58,
                   C["bg"], 0.70)

        # Main alarm text
        sev_text = f"[ {severity} ALERT ]"
        (tw, _), _ = cv2.getTextSize(sev_text, cv2.FONT_HERSHEY_DUPLEX, 1.2, 2)
        tx = (feed_w - tw) // 2
        if int(t * 2) % 2 == 0:
            cv2.putText(frame, sev_text, (tx, fh_cy - 8),
                        cv2.FONT_HERSHEY_DUPLEX, 1.2,
                        sev_col(severity), 2, cv2.LINE_AA)

        # Sub text
        count_line = f"{person_count} PERSON{'S' if person_count != 1 else ''}" \
                     f"  ·  {get_threat(person_count)}"
        (tw2, _), _ = cv2.getTextSize(count_line, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 1)
        put_c(frame, count_line, feed_w // 2, fh_cy + 34,
              0.65, C["orange"], 1)

        # Alarm ID bottom-left of feed
        put(frame, f"◈  ALARM ID: ALM-{alarm_count:04d}", 14, h - 68,
            0.40, C["orange"])

        # Animated corner inward sweeps
        sweep = int((t * 80) % 40)
        line(frame, (0, 85), (sweep, 85), sev_col(severity), 2)
        line(frame, (0, 85), (0, 85 + sweep), sev_col(severity), 2)
        line(frame, (feed_w, 85), (feed_w - sweep, 85), sev_col(severity), 2)
        line(frame, (feed_w, 85), (feed_w, 85 + sweep), sev_col(severity), 2)

# ═══════════════════════════════════════════════════════════════════════════════
#  SNAPSHOT  — auto-saved to snapshots/ on every alarm trigger
# ═══════════════════════════════════════════════════════════════════════════════
def save_snapshot(frame, alarm_no):
    os.makedirs(SNAPSHOT_DIR, exist_ok=True)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = os.path.join(SNAPSHOT_DIR, f"ALM-{alarm_no:04d}_{ts}.jpg")
    cv2.imwrite(fname, frame)
    print(f"  [SNAPSHOT] Saved: {fname}")
    return fname

# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL LOG
# ═══════════════════════════════════════════════════════════════════════════════
def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)
def _font(size=10, bold=False, color="1C1C1E"):
    return Font(name="Arial", size=size, bold=bold, color=color)
def _border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)
def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

HEADERS = [
    ("Alarm ID",12),("Date",12),("Time",10),("Site ID",22),
    ("Site Name",22),("Location",20),("Zone",32),("Camera ID",14),
    ("Persons Detected",10),("Threat Level",20),("Confidence (%)",12),
    ("Severity",12),("Alarm Status",16),("Duration (sec)",12),
    ("Response Required",38),("Operator",18),("Snapshot",30),("Remarks",40),
]

def init_log():
    if os.path.exists(LOG_FILE): return
    wb = Workbook()
    ws = wb.active
    ws.title = "Alarm Log"
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    tc = ws["A1"]
    tc.value     = f"AIRTEL BTS TACTICAL MONITORING — {SITE_NAME} ({SITE_ID}) — {SITE_LOCATION}"
    tc.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    tc.fill      = _fill("D0021B")
    tc.alignment = _align("center")
    ws.row_dimensions[1].height = 28
    for col, (label, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill      = _fill("1A1A2E")
        c.alignment = _align("center")
        c.border    = Border(bottom=Side(style="medium", color="0066CC"),
                             left=Side(style="thin",   color="333333"),
                             right=Side(style="thin",  color="333333"))
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}2"
    wb.save(LOG_FILE)

def log_event(alarm_no, count, conf, severity, status, duration, remarks="", snapshot=""):
    now = datetime.now()
    wb  = load_workbook(LOG_FILE)
    ws  = wb["Alarm Log"]
    row = ws.max_row + 1
    if   status == "ALARM RAISED"  and severity == "CRITICAL": row_bg = "FFF0F0"
    elif status == "ALARM RAISED":                              row_bg = "FFF8F0"
    elif status == "ALARM CLEARED":                             row_bg = "F0FFF4"
    else: row_bg = "F5F5F7" if row % 2 == 0 else "FFFFFF"

    values = [
        f"ALM-{alarm_no:04d}", now.strftime("%d-%m-%Y"), now.strftime("%H:%M:%S"),
        SITE_ID, SITE_NAME, SITE_LOCATION, ZONE, CAMERA_ID, count,
        get_threat(count) if count > 0 else "CLEAR",
        round(conf * 100, 1) if conf > 0 else 0.0,
        severity, status, round(duration, 1),
        get_response(count) if count > 0 else "NO ACTION REQUIRED",
        OPERATOR, snapshot, remarks,
    ]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(row_bg); c.border = _border()
        if col == 12:
            c.font = Font(name="Arial", size=10, bold=True,
                          color="C0392B" if val=="CRITICAL" else "D35400")
            c.fill = _fill("FFE5E5" if val=="CRITICAL" else "FFF3E5")
            c.alignment = _align("center")
        elif col == 13:
            cols = {"ALARM RAISED":"C0392B","ALARM ACTIVE":"D35400",
                    "ALARM CLEARED":"1A7A3C","SESSION ENDED":"555555"}
            c.font = Font(name="Arial", size=9, bold=True,
                          color=cols.get(val,"333333"))
            c.alignment = _align("center")
        elif col in (1,2,3,8,9,11,14):
            c.font = _font(bold=(col==1)); c.alignment = _align("center")
        else:
            c.font = _font(); c.alignment = _align(wrap=(col in (7,15,17)))
    ws.row_dimensions[row].height = 20
    wb.save(LOG_FILE)

# ═══════════════════════════════════════════════════════════════════════════════
#  BOOT SCREEN
# ═══════════════════════════════════════════════════════════════════════════════
def draw_boot(frame, progress, msg):
    h, w = frame.shape[:2]
    rect(frame, (0,0), (w,h), C["bg"])
    cx, cy = w//2, h//2
    # Outer rings
    for r, a in [(120,0.15),(90,0.25),(60,0.40)]:
        draw_arc_ring(frame, cx, cy, r, 0, 360, C["border2"])
    # Rotating arc
    ang = (progress * 360 * 3) % 360
    draw_arc_ring(frame, cx, cy, 90, int(ang), int(ang)+120, C["cyan"], 2)
    draw_arc_ring(frame, cx, cy, 60, int(ang)+180, int(ang)+260, C["orange"], 2)
    # Center logo
    put_c(frame, "AIRTEL", cx, cy - 12, 0.80, C["airtel"], 2,)
    put_c(frame, "TACTICAL SURVEILLANCE", cx, cy + 16, 0.38, C["cyan_dim"])
    # Progress bar
    bw = 300
    bx = cx - bw//2
    rect(frame, (bx, cy+50), (bx+bw, cy+64), C["border"])
    fill_w = int(bw * progress)
    rect(frame, (bx, cy+50), (bx+fill_w, cy+64), C["cyan"])
    put_c(frame, f"{int(progress*100)}%", cx, cy+80, 0.38, C["gray"])
    put_c(frame, msg, cx, cy+100, 0.34, C["cyan_dim"])
    # Scanlines
    scanline_overlay(frame, 0.12)

# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print("\n" + "=" * 68)
    print("  AIRTEL BTS TACTICAL SURVEILLANCE — IRON MAN HUD EDITION")
    print("=" * 68)
    print(f"  Site     : {SITE_NAME} ({SITE_ID})")
    print(f"  Location : {SITE_LOCATION}")
    print(f"  Camera   : {CAMERA_ID}  [source={CAMERA_SOURCE}]")
    print(f"  Model    : {MODEL_PATH}  (upgrade: set MODEL_PATH='yolov8m.pt')")
    print(f"  Trigger  : {CONSECUTIVE_REQUIRED} consecutive frames @ {CONFIDENCE_THRESHOLD:.0%}+ conf")
    print(f"  Log      : {LOG_FILE}")
    print("=" * 68)
    print("  Press Q to quit  |  Press S to screenshot\n")

    init_log()

    # ── Boot animation ────────────────────────────────────────────────────────
    boot_frame = np.zeros((720, 1280, 3), dtype=np.uint8)
    cv2.imshow("Airtel BTS — Tactical Surveillance", boot_frame)
    boot_msgs = [
        "INITIALISING SUBSYSTEMS...",
        "LOADING AI MODEL...",
        "CALIBRATING SENSOR ARRAY...",
        "ESTABLISHING SECURE LINK...",
        "TACTICAL HUD ONLINE",
    ]
    for i, msg in enumerate(boot_msgs):
        for p in range(20):
            prog = (i * 20 + p) / 100.0
            draw_boot(boot_frame, prog, msg)
            cv2.imshow("Airtel BTS — Tactical Surveillance", boot_frame)
            if cv2.waitKey(20) & 0xFF == ord("q"):
                cv2.destroyAllWindows(); return

    print("Loading YOLOv8 model...")
    model = YOLO(MODEL_PATH)
    model.overrides['iou'] = NMS_IOU_THRESHOLD
    print("Model ready.\n")

    # Finish boot
    draw_boot(boot_frame, 1.0, "ALL SYSTEMS NOMINAL — COMMENCING SURVEILLANCE")
    cv2.imshow("Airtel BTS — Tactical Surveillance", boot_frame)
    cv2.waitKey(800)

    cap = cv2.VideoCapture(CAMERA_SOURCE)
    if not cap.isOpened():
        print(f"ERROR: Cannot open camera {CAMERA_SOURCE}"); return

    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    cap.set(cv2.CAP_PROP_BUFFERSIZE,   1)        # minimise latency

    consecutive       = 0
    alarming          = False
    alarm_count       = 0
    alarm_start       = None
    last_alarm_logged = 0
    last_cleared_ts   = None
    fps               = 0.0
    prev_time         = time.time()
    severity          = "HIGH"
    session_start     = time.time()
    screenshot_count  = 0

    print(f"{'Time':10}  {'Alarm ID':10}  {'Event':<30}  {'#':>3}  {'Conf':>6}  Severity")
    print("-" * 74)

    while True:
        ret, frame = cap.read()
        if not ret:
            time.sleep(0.3); continue

        now = time.time()
        t   = now                       # time reference for animations
        fps = 0.9 * fps + 0.1 * (1.0 / max(now - prev_time, 1e-6))
        prev_time = now
        # ── Inference ─────────────────────────────────────────────────────────
        results    = model(frame, verbose=False,
                           classes=[0,15,16,17,18,19,20,21,22,23],
                           conf=CONFIDENCE_THRESHOLD,
                           iou=NMS_IOU_THRESHOLD)[0]
        detections = []
        for box in results.boxes:
            conf_v = float(box.conf[0])
            cls    = int(box.cls[0])
            if conf_v >= CONFIDENCE_THRESHOLD:
                x1, y1, x2, y2 = [int(v) for v in box.xyxy[0]]
                lbl = "Person" if cls == 0 else ANIMAL_CLASSES.get(cls, "Animal")
                detections.append((conf_v, x1, y1, x2, y2, lbl))

        detections.sort(key=lambda d: d[0], reverse=True)
        person_count = sum(1 for d in detections if d[5] == "Person")
        best_conf    = detections[0][0] if detections else 0.0
        severity     = get_severity(person_count) if person_count > 0 else "CLEAR"

        # ── Draw target boxes ─────────────────────────────────────────────────
        for idx, (conf_v, x1, y1, x2, y2, lbl) in enumerate(detections, 1):
            draw_target_box(frame, x1, y1, x2, y2, conf_v, idx, severity, lbl, t)

        # ── Consecutive window ────────────────────────────────────────────────
        if person_count > 0:
            consecutive = min(consecutive + 1, CONSECUTIVE_REQUIRED)
        else:
            consecutive = max(consecutive - 1, 0)

        ts = datetime.now().strftime("%H:%M:%S")

        # ── Alarm state machine ───────────────────────────────────────────────
        if consecutive >= CONSECUTIVE_REQUIRED:
            if not alarming:
                alarming = True; alarm_count += 1
                alarm_start = now; last_alarm_logged = 0
                play_alarm(severity)
                snap = save_snapshot(frame, alarm_count)
                alm_id = f"ALM-{alarm_count:04d}"
                print(f"{ts:10}  {alm_id:10}  {'INTRUSION ALARM RAISED':<30}  "
                      f"{person_count:>3}  {best_conf:>5.0%}  {severity}")
                log_event(alarm_count, person_count, best_conf, severity,
                          "ALARM RAISED", 0.0,
                          f"{get_threat(person_count)} — alarm triggered",
                          snapshot=snap)
            elif now - last_alarm_logged >= ALARM_COOLDOWN_SEC:
                dur = now - alarm_start
                log_event(alarm_count, person_count, best_conf, severity,
                          "ALARM ACTIVE", dur,
                          f"Ongoing — {dur:.0f}s elapsed")
                last_alarm_logged = now
        else:
            if alarming:
                dur = now - alarm_start if alarm_start else 0
                last_cleared_ts = ts
                print(f"{ts:10}  ALM-{alarm_count:04d}  {'ALARM CLEARED — AREA SECURE':<30}  "
                      f"{'0':>3}  {'—':>6}  {dur:.1f}s")
                log_event(alarm_count, 0, 0, severity,
                          "ALARM CLEARED", dur,
                          f"Area clear after {dur:.1f}s")
            alarming = False; alarm_start = None; last_alarm_logged = 0

        # ── Draw HUD ──────────────────────────────────────────────────────────
        draw_hud(frame, fps, consecutive, alarming, alarm_count,
                 person_count, best_conf, severity, alarm_start,
                 last_cleared_ts, detections, session_start, t)

        cv2.imshow("Airtel BTS — Tactical Surveillance", frame)
        key = cv2.waitKey(1) & 0xFF
        if key == ord("q"):
            break
        if key == ord("s"):
            fname = f"screenshot_{screenshot_count:03d}.png"
            cv2.imwrite(fname, frame)
            screenshot_count += 1
            print(f"Screenshot saved: {fname}")

    # ── Session end ───────────────────────────────────────────────────────────
    if alarming and alarm_start:
        dur = time.time() - alarm_start
        log_event(alarm_count, 0, 0, severity, "SESSION ENDED", dur,
                  "Session ended by operator")

    cap.release()
    cv2.destroyAllWindows()
    session_dur = int(time.time() - session_start)
    h_s, m_s, s_s = session_dur//3600, (session_dur%3600)//60, session_dur%60
    snaps = len(os.listdir(SNAPSHOT_DIR)) if os.path.exists(SNAPSHOT_DIR) else 0
    print("\n" + "=" * 68)
    print(f"  SESSION SUMMARY")
    print(f"  Duration       : {h_s:02d}:{m_s:02d}:{s_s:02d}")
    print(f"  Total alarms   : {alarm_count}")
    print(f"  Snapshots saved: {snaps}  ({SNAPSHOT_DIR}/)")
    print(f"  Excel log      : {os.path.abspath(LOG_FILE)}")
    print("=" * 68 + "\n")


if __name__ == "__main__":
    main()
    #iwalkalonelyroad
    